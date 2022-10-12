import openpyxl as xl
from openpyxl.utils import get_column_letter

#Import excel document
wb = xl.load_workbook('menu.xlsx')
sheet = wb['Menu']

#from
fromArr = [
    "Pomegranate-Berry Burst 100 tablets", 
    "Lemon-Lime Blast 100 tablets", 
    "Ignite-Me Orange 100 tablets", 
    "Tropical Fruit Force 100 tablets",
    "Cranberry 1/2 Gallon",
    "Mango 1/2 Gallon",
    "Mandarin Half Gallon",
    "Raspberry 1.8 Oz",
    "Lemon 1.8 Oz",
    "Original 1.8 Oz",
    "Peach 1.8 Oz",
    "Açaí Berry 5 lb. Pouch",
    "Banana 5 lb",
    "Blackberry 5 lb",
    "Blue Blast 5 lb",
    "Blue Raspberry 5 lb",
    "Cherry 5 lb",
    "Cherry Limeade 5 lb",
    "Coconut 5 lb",
    "Cranberry 5 lb",
    "Cucumber Lime 5 lb",
    "Dragon Fruit 5 lb",
    "Grape 5 lb",
    "Green Apple 5 lb",
    "Lemonade 5 lb",
    "Limeade 5 lb",
    "Mango 5 lb",
    "Margarita 5 lb",
    "Melon 5 lb",
    "Orange 5 lb",
    "Orange Pineapple 5 lb",
    "Passion Fruit 5 lb",
    "Peach 5 lb",
    "Piña Colada 5 lb",
    "Pineapple 5 lb",
    "Pineapple Fandango 5 lb",
    "Pink Lemonade 5 lb",
    "Pomegranate 5 lb",
    "Prickly Pear 5 lb",
    "Rainbow Candy 5 lb",
    "Raspberry 5 lb",
    "Raspberry Lemonade 5 lb",
    "Sour Apple 5 lb",
    "Strawberry 5 lb",
    "Tropical Fruit Punch 5 lb",
    "Watermelon 5 lb",
    "White Grape 5 lb",
]

#to

toArr = [
    {'sku': '081K', 'name': 'Liftoff® Pomegranate-Berry Burst 30 Tablets', 'serving': 1},
    {'sku': '3152', 'name': 'Liftoff® Lemon-Lime Blast 10 Tablets', 'serving': 1},
    {'sku': '3277', 'name': 'Liftoff® Ignite-Me Orange 30 Tablets', 'serving': 1},
    {'sku': '2696', 'name': 'Liftoff® Tropical Fruit Force 10 Tablets', 'serving': 1},
    {'sku': '1189', 'name': 'Herbal Aloe Concentrate: Cranberry Pint', 'serving': 1},
    {'sku': '1065', 'name': 'Herbal Aloe Concentrate: Mango Pint', 'serving': 1},
    {'sku': '2631', 'name': 'Herbal Aloe Concentrate: Mandarin Pint', 'serving': 1},
    {'sku': '0189', 'name': 'Herbal Tea Concentrate: Raspberry 3.6 OZ (102g)', 'serving': 1.5},
    {'sku': '0188', 'name': 'Herbal Tea Concentrate: Lemon 3.6 OZ (102g)', 'serving': 1.5},
    {'sku': '0106', 'name': 'Herbal Tea Concentrate: Original 3.6 OZ (102g)', 'serving': 1.5},
    {'sku': '0190', 'name': 'Herbal Tea Concentrate: Peach 3.6 OZ (102g)', 'serving': 1.5},
    {'sku': '218M', 'name': 'Drink Mix: Açaí Berry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '686M', 'name': 'Drink Mix: Banana 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '218M', 'name': 'Drink Mix: Açaí Berry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '176M', 'name': 'Drink Mix: Blackberry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '177M', 'name': 'Drink Mix: Blue Blast 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '178M', 'name': 'Drink Mix: Blue Raspberry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '180M', 'name': 'Drink Mix: Cherry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '181M', 'name': 'Drink Mix: Cherry Limeade 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '182M', 'name': 'Drink Mix: Coconut 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '185M', 'name': 'Drink Mix: Cranberry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '183M', 'name': 'Drink Mix: Cucumber Lime 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '187M', 'name': 'Drink Mix: Dragon Fruit 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '188M', 'name': 'Drink Mix: Grape 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '189M', 'name': 'Drink Mix: Green Apple 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '195M', 'name': 'Drink Mix: Lemonade 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '196M', 'name': 'Drink Mix: Limeade 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '219M', 'name': 'Drink Mix: Mango 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '220M', 'name': 'Drink Mix: Margarita 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '221M', 'name': 'Drink Mix: Melon 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '197M', 'name': 'Drink Mix: Orange 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '198M', 'name': 'Drink Mix: Orange Pineapple 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '199M', 'name': 'Drink Mix: Passion Fruit 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '200M', 'name': 'Drink Mix: Peach 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '202M', 'name': 'Drink Mix: Piña Colada 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '203M', 'name': 'Drink Mix: Pineapple 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '204M', 'name': 'Drink Mix: Pineapple Fandago 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '206M', 'name': 'Drink Mix: Pink Lemonade 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '207M', 'name': 'Drink Mix: Pomegranate 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '222M', 'name': 'Drink Mix: Prickly Pear 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '208M', 'name': 'Drink Mix: Rainbow Candy 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '223M', 'name': 'Drink Mix: Raspberry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '224M', 'name': 'Drink Mix: Raspberry Lemonade 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '209M', 'name': 'Drink Mix: Sour Apple 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '211M', 'name': 'Drink Mix: Strawberry 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '213M', 'name': 'Drink Mix: Tropical Fruit Punch 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '214M', 'name': 'Drink Mix: Watermelon 4.4 Oz. Pouch', 'serving': 1},
    {'sku': '216M', 'name': 'Drink Mix: White Grape 4.4 Oz. Pouch', 'serving': 1}

    
]

#loop through the table
count = 0
for row in range(3, sheet.max_row + 1):
    sku = sheet.cell(row, 10).value
    name = sheet.cell(row, 11).value
    serving = sheet.cell(row,12).value
    
    for item in range(len(fromArr) - 1) :
        if (fromArr[item] in name):
            newname = name.replace(name,toArr[item]['sku'])
            newsku = toArr[item]['name']
            newserving = toArr[item]['serving']
            
            _ = sheet.cell(column=11, row=row, value=newsku)
            _ = sheet.cell(column=10, row=row, value=newname)
            _ = sheet.cell(column=12, row=row, value=newserving)
            count+=1

        
print("Made " + str(count) + " changes.")
wb.save(filename='updated_menu.xlsx')
